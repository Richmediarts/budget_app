from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, Response
from datetime import datetime, timedelta
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import calendar
import csv
import io
import json
import os
import database as db

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY') or 'budget-app-secret-key-2024'

# Plaid configuration — from plaid_config.json or env vars
CONFIG_DIR = os.path.dirname(os.path.abspath(__file__))
plaid_cfg_file = os.path.join(CONFIG_DIR, 'plaid_config.json')
if os.path.exists(plaid_cfg_file):
    with open(plaid_cfg_file) as f:
        plaid_cfg = json.load(f)
else:
    plaid_cfg = {}
PLAID_CLIENT_ID = plaid_cfg.get('client_id') or os.environ.get('PLAID_CLIENT_ID', '')
PLAID_SECRET = plaid_cfg.get('secret') or os.environ.get('PLAID_SECRET', '')
PLAID_ENV = plaid_cfg.get('environment') or os.environ.get('PLAID_ENV', 'sandbox')
PLAID_PRODUCTS = ['transactions']
PLAID_COUNTRY_CODES = ['US']

db.init_db()

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

@app.before_request
def check_auth():
    if request.endpoint in ('login', 'logout', 'setup', 'static'):
        return
    if 'user_id' not in session:
        return redirect(url_for('login'))

def add_months(dt, months):
    # Safely add months to a date
    year = dt.year + (dt.month - 1 + months) // 12
    month = (dt.month - 1 + months) % 12 + 1
    day = min(dt.day, calendar.monthrange(year, month)[1])
    return dt.replace(year=year, month=month, day=day)

def get_category_options():
    hierarchical = db.get_budget_categories()
    options = []
    all_limits = 0
    for parent in hierarchical:
        options.append({'id': parent['id'], 'display': parent['name']})
        all_limits += parent['monthly_limit']
        for child in parent.get('children', []):
            options.append({'id': child['id'], 'display': '  └─ ' + child['name']})
            all_limits += child['monthly_limit']
    db._cached_total_budget = all_limits
    return options

def get_hierarchical_and_flat():
    hierarchical = db.get_budget_categories()
    flat = []
    for parent in hierarchical:
        flat.append(parent)
        flat.extend(parent.get('children', []))
    return hierarchical, flat

@app.context_processor
def inject_globals():
    cat_opts = get_category_options()
    bank_accounts = db.get_all_bank_accounts()
    payees = db.get_all_payees()

    logo_map = {
        'first tech federal credit union': {'small': 'firsttech_circle_logo.png', 'bg': 'firsttech_logo_bg.png'},
        'pnc': {'small': 'pnc_circle_logo.png', 'bg': 'pnc_bg.png'},
        'navy federal credit union': {'small': 'navy_federal_circle_logo.png', 'bg': None},
        'capital one': {'small': 'capital_one_circle_logo.png', 'bg': None},
        'bridgecrest': {'small': 'bridgecrest_circle_logo.png', 'bg': None},
    }

    account_type_options = [
        ('checking', 'Checking'),
        ('savings', 'Savings'),
        ('money_market', 'Money Market'),
        ('cd', 'CD'),
        ('investment', 'Investment'),
        ('loan', 'Loan'),
    ]

    return {
        'current_date': datetime.now().strftime('%B %d, %Y'),
        'datetime': datetime,
        'total_balance': sum(a['current_balance'] for a in bank_accounts),
        'total_debt': sum(c['current_balance'] for c in db.get_all_credit_cards()),
        'total_budget': db._cached_total_budget,
        'next_paycheck_date': db.get_next_paycheck_date(),
        'category_options': cat_opts,
        'bank_accounts': bank_accounts,
        'payees': payees,
        'logo_map': logo_map,
        'account_type_options': account_type_options,
    }

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        user = db.get_user_by_username(username)
        if user and check_password_hash(user['password_hash'], password):
            session['user_id'] = user['id']
            session['username'] = user['username']
            flash('Logged in successfully.', 'success')
            return redirect(url_for('dashboard'))
        flash('Invalid username or password.', 'danger')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('Logged out.', 'info')
    return redirect(url_for('login'))

@app.route('/setup', methods=['GET', 'POST'])
def setup():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    if db.has_users():
        return redirect(url_for('login'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        confirm = request.form.get('confirm_password', '')
        if not username or not password:
            flash('Username and password are required.', 'danger')
        elif password != confirm:
            flash('Passwords do not match.', 'danger')
        elif len(password) < 6:
            flash('Password must be at least 6 characters.', 'danger')
        else:
            password_hash = generate_password_hash(password)
            if db.create_user(username, password_hash):
                flash('Account created! Please log in.', 'success')
                return redirect(url_for('login'))
            flash('Username already taken.', 'danger')
    return render_template('setup.html')

@app.route('/')
def dashboard():
    stats = db.get_dashboard_stats()
    return render_template('dashboard.html', stats=stats, next_paycheck_date=db.get_next_paycheck_date())

@app.route('/paychecks')
def paychecks():
    paychecks = db.get_all_paychecks()
    next_paycheck = db.get_next_paycheck_date()
    today = datetime.now().strftime('%Y-%m-%d')
    return render_template('paychecks.html', paychecks=paychecks, next_paycheck=next_paycheck, today=today, active_tab='history')

@app.route('/api/payslip/<int:paycheck_id>')
def get_payslip(paycheck_id):
    paychecks = db.get_all_paychecks()
    pc = next((p for p in paychecks if p['id'] == paycheck_id), None)
    if pc:
        return jsonify(pc)
    return jsonify({'error': 'Paycheck not found'}), 404

@app.route('/add_paycheck', methods=['GET'])
def add_paycheck_form():
    import json
    prefill = {}
    if request.args.get('prefill'):
        try:
            prefill = json.loads(request.args.get('prefill'))
        except:
            pass
    return render_template('add_paycheck.html', paycheck=prefill)


@app.route('/add_paycheck', methods=['POST'])
def add_paycheck():
    db.add_paycheck(
        pay_date=request.form.get('check_date', ''),
        pay_period_begin=request.form.get('pay_period_begin', ''),
        pay_period_end=request.form.get('pay_period_end', ''),
        check_date=request.form.get('check_date', ''),
        check_number=request.form.get('check_number', ''),
        employee_name=request.form.get('employee_name', ''),
        employee_id=request.form.get('employee_id', ''),
        company=request.form.get('company', ''),
        hours_worked=float(request.form.get('hours_worked') or 0),
        gross_pay=float(request.form.get('gross_pay') or 0),
        pre_tax_deductions=float(request.form.get('pre_tax_deductions') or 0),
        employee_taxes=float(request.form.get('employee_taxes') or 0),
        post_tax_deductions=float(request.form.get('post_tax_deductions') or 0),
        net_pay=float(request.form.get('net_pay') or 0),
        salary=float(request.form.get('salary') or 0),
        biometric_credit=float(request.form.get('biometric_credit') or 0),
        floating_holiday=float(request.form.get('floating_holiday') or 0),
        holiday_pay=float(request.form.get('holiday_pay') or 0),
        vacation_pay=float(request.form.get('vacation_pay') or 0),
        group_term_life=float(request.form.get('group_term_life') or 0),
        spousal_biometric=float(request.form.get('spousal_biometric') or 0),
        other_earnings=float(request.form.get('other_earnings') or 0),
        oasdi=float(request.form.get('oasdi') or 0),
        medicare=float(request.form.get('medicare') or 0),
        federal_tax=float(request.form.get('federal_tax') or 0),
        state_tax=float(request.form.get('state_tax') or 0),
        state_name=request.form.get('state_name', ''),
        social_security=float(request.form.get('social_security') or 0),
        retirement_401k=float(request.form.get('retirement_401k') or 0),
        add_insurance=float(request.form.get('add_insurance') or 0),
        dental_plan=float(request.form.get('dental_plan') or 0),
        eye_plan=float(request.form.get('eye_plan') or 0),
        health_care_fsa=float(request.form.get('health_care_fsa') or 0),
        health_insurance=float(request.form.get('health_insurance') or 0),
        optional_life=float(request.form.get('optional_life') or 0),
        hsa=float(request.form.get('hsa') or 0),
        loan_repayment=float(request.form.get('loan_repayment') or 0),
        dependent_life=float(request.form.get('dependent_life') or 0),
        stock_purchase=float(request.form.get('stock_purchase') or 0),
        spousal_life=float(request.form.get('spousal_life') or 0),
        employer_match=float(request.form.get('employer_match') or 0),
        employer_hsa=float(request.form.get('employer_hsa') or 0),
        federal_filing_status=request.form.get('federal_filing_status', ''),
        state_filing_status=request.form.get('state_filing_status', ''),
        bank_name=request.form.get('bank_name', ''),
        account_number=request.form.get('account_number', ''),
        deposit_amount=float(request.form.get('deposit_amount') or 0),
        bank2_name=request.form.get('bank2_name', ''),
        account2_number=request.form.get('account2_number', ''),
        deposit2_amount=float(request.form.get('deposit2_amount') or 0),
        gross_pay_ytd=float(request.form.get('gross_pay_ytd') or 0),
        pre_tax_deductions_ytd=float(request.form.get('pre_tax_deductions_ytd') or 0),
        employee_taxes_ytd=float(request.form.get('employee_taxes_ytd') or 0),
        post_tax_deductions_ytd=float(request.form.get('post_tax_deductions_ytd') or 0),
        net_pay_ytd=float(request.form.get('net_pay_ytd') or 0),
        notes=request.form.get('notes', '')
    )
    flash('Paycheck added successfully!', 'success')
    return redirect(url_for('paychecks'))

@app.route('/delete_paycheck/<int:id>')
def delete_paycheck(id):
    db.delete_paycheck(id)
    flash('Paycheck deleted.', 'info')
    return redirect(url_for('paychecks'))

@app.route('/edit_paycheck/<int:id>')
def edit_paycheck(id):
    paycheck = db.get_paycheck(id)
    if not paycheck:
        flash('Paycheck not found.', 'danger')
        return redirect(url_for('paychecks'))
    paychecks = db.get_all_paychecks()
    next_paycheck = db.get_next_paycheck_date()
    today = datetime.now().strftime('%Y-%m-%d')
    return render_template('paychecks.html', paychecks=paychecks, next_paycheck=next_paycheck, today=today, edit_paycheck=paycheck, active_tab='history')

@app.route('/update_paycheck/<int:id>', methods=['POST'])
def update_paycheck(id):
    numeric_fields = ['hours_worked', 'gross_pay', 'net_pay', 'pre_tax_deductions', 'employee_taxes', 
                      'post_tax_deductions', 'salary', 'biometric_credit', 'floating_holiday', 
                      'holiday_pay', 'vacation_pay', 'group_term_life', 'spousal_biometric',
                      'oasdi', 'medicare', 'federal_tax', 'state_tax', 'social_security',
                      'retirement_401k', 'add_insurance', 'dental_plan', 'eye_plan', 
                      'health_care_fsa', 'health_insurance', 'optional_life', 'hsa',
                      'loan_repayment', 'dependent_life', 'stock_purchase', 'spousal_life',
                      'employer_match', 'employer_hsa', 'deposit_amount', 'deposit2_amount',
                      'gross_pay_ytd', 'pre_tax_deductions_ytd', 'employee_taxes_ytd',
                      'post_tax_deductions_ytd', 'net_pay_ytd']
    
    kwargs = {'id': id}
    for field in numeric_fields:
        val = request.form.get(field, '0').replace(',', '')
        kwargs[field] = float(val) if val else 0
    
    kwargs['pay_date'] = request.form.get('check_date', '')
    kwargs['pay_period_begin'] = request.form.get('pay_period_begin', '')
    kwargs['pay_period_end'] = request.form.get('pay_period_end', '')
    kwargs['check_date'] = request.form.get('check_date', '')
    kwargs['check_number'] = request.form.get('check_number', '')
    kwargs['employee_name'] = request.form.get('employee_name', '')
    kwargs['employee_id'] = request.form.get('employee_id', '')
    kwargs['company'] = request.form.get('company', '')
    kwargs['state_name'] = request.form.get('state_name', '')
    kwargs['federal_filing_status'] = request.form.get('federal_filing_status', '')
    kwargs['state_filing_status'] = request.form.get('state_filing_status', '')
    kwargs['bank_name'] = request.form.get('bank_name', '')
    kwargs['account_number'] = request.form.get('account_number', '')
    kwargs['bank2_name'] = request.form.get('bank2_name', '')
    kwargs['account2_number'] = request.form.get('account2_number', '')
    kwargs['notes'] = request.form.get('notes', '')
    
    db.update_paycheck(id, **kwargs)
    flash('Paycheck updated successfully!', 'success')
    return redirect(url_for('paychecks'))

import re
from dateutil import parser as date_parser

def parse_paycheck_text(raw_text):
    result = {}
    lines = raw_text.strip().split('\n')
    text = raw_text
    
    def extract_money(s):
        s = s.strip().replace('$', '').replace(',', '')
        try:
            return float(s)
        except:
            return None
    
    def parse_date(s):
        s = s.strip()
        for fmt in ['%m/%d/%Y', '%Y-%m-%d', '%m-%d-%Y', '%B %d, %Y']:
            try:
                return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
            except:
                pass
        try:
            return date_parser.parse(s).strftime('%Y-%m-%d')
        except:
            return None
    
    for i, line in enumerate(lines):
        line_lower = line.lower()
        
        if 'voyix' in line_lower:
            result['company'] = 'NCR Voyix Corporation'
        
        emp_id = re.search(r'\b(\d{10,11})\b', line)
        if emp_id and 'employee_id' not in result:
            result['employee_id'] = emp_id.group(1)
        
        dates = re.findall(r'(\d{2}/\d{2}/\d{4})', line)
        if len(dates) >= 2:
            begin_date = parse_date(dates[0])
            end_date = parse_date(dates[1])
            if begin_date and 'pay_period_begin' not in result:
                result['pay_period_begin'] = begin_date
            if end_date and 'pay_period_end' not in result:
                result['pay_period_end'] = end_date
        
        if len(dates) >= 3 and 'check_date' not in result:
            cd = parse_date(dates[2])
            if cd:
                result['check_date'] = cd
    
    for i, line in enumerate(lines):
        line_lower = line.lower()
        words = line.split()
        
        if 'richard' in line_lower or 'johnson' in line_lower:
            name_words = []
            for word in words:
                if word[0].isupper() and len(word) > 2 and not any(c.isdigit() for c in word):
                    if 'voyix' not in word.lower() and 'corporation' not in word.lower() and word not in ['Name', 'Company', 'Employee', 'Description', 'Amount', 'Current', 'YTD', 'Hours', 'Gross', 'Tax', 'Net']:
                        name_words.append(word)
                        if len(name_words) >= 2:
                            break
            if name_words:
                result['employee_name'] = ' '.join(name_words)
    
    summary_line = None
    for line in lines:
        line_lower = line.lower()
        if 'current' in line_lower and re.search(r'\b\d{2}\.\d{2}\b', line) and not line_lower.strip().startswith('hours worked'):
            money_matches = re.findall(r'([\d,]+\.\d{2})', line)
            if len(money_matches) >= 5:
                summary_line = line
                break
    
    if summary_line:
        money_matches = re.findall(r'([\d,]+\.\d{2})', summary_line)
        if len(money_matches) >= 6:
            result['hours_worked'] = extract_money(money_matches[0])
            result['gross_pay'] = extract_money(money_matches[1])
            result['pre_tax_deductions'] = extract_money(money_matches[2])
            result['employee_taxes'] = extract_money(money_matches[3])
            result['post_tax_deductions'] = extract_money(money_matches[4])
            result['net_pay'] = extract_money(money_matches[5])
    
    ytd_line = None
    for line in lines:
        line_lower = line.lower()
        if 'ytd' in line_lower and re.search(r'^\s*YTD', line):
            money_matches = re.findall(r'([\d,]+\.\d{2})', line)
            if len(money_matches) >= 5:
                ytd_line = line
                break
    
    if ytd_line:
        money_matches = re.findall(r'([\d,]+\.\d{2})', ytd_line)
        if len(money_matches) >= 6:
            result['hours_worked_ytd'] = extract_money(money_matches[0])
            result['gross_pay_ytd'] = extract_money(money_matches[1])
            result['pre_tax_deductions_ytd'] = extract_money(money_matches[2])
            result['employee_taxes_ytd'] = extract_money(money_matches[3])
            result['post_tax_deductions_ytd'] = extract_money(money_matches[4])
            result['net_pay_ytd'] = extract_money(money_matches[5])
    
    def get_value_after_label(line, label, default=None, ytd=False):
        idx = line.lower().find(label)
        if idx == -1:
            return default
        after_label = line[idx + len(label):]
        matches = re.findall(r'([\d,]+\.\d{2})', after_label)
        if matches:
            idx = 1 if ytd else 0
            if idx < len(matches):
                return extract_money(matches[idx])
            return extract_money(matches[0])
        return default
    
    for i, line in enumerate(lines):
        line_lower = line.lower()
        
        if any(x in line_lower for x in ['401k savings plan', '401(k) savings plan']):
            result['retirement_401k'] = get_value_after_label(line, '401k')
            result['retirement_401k_ytd'] = get_value_after_label(line, '401k', ytd=True)
        
        if 'medical' in line_lower and ('plan' in line_lower or 'ins' in line_lower):
            result['health_insurance'] = get_value_after_label(line, 'medical')
            result['health_insurance_ytd'] = get_value_after_label(line, 'medical', ytd=True)
        
        if 'dental plan' in line_lower:
            result['dental_plan'] = get_value_after_label(line, 'dental')
            result['dental_plan_ytd'] = get_value_after_label(line, 'dental', ytd=True)
        
        if 'eye plan' in line_lower:
            result['eye_plan'] = get_value_after_label(line, 'eye')
            result['eye_plan_ytd'] = get_value_after_label(line, 'eye', ytd=True)
        
        if 'health care fsa' in line_lower:
            result['health_care_fsa'] = get_value_after_label(line, 'fsa')
            result['health_care_fsa_ytd'] = get_value_after_label(line, 'fsa', ytd=True)
        
        if 'optional life' in line_lower:
            result['optional_life'] = get_value_after_label(line, 'optional life')
            result['optional_life_ytd'] = get_value_after_label(line, 'optional life', ytd=True)
        
        if 'add insurance' in line_lower:
            result['add_insurance'] = get_value_after_label(line, 'add')
            result['add_insurance_ytd'] = get_value_after_label(line, 'add', ytd=True)
        
        if 'federal withholding' in line_lower and 'taxable' not in line_lower:
            result['federal_tax'] = get_value_after_label(line, 'federal withholding')
            result['federal_tax_ytd'] = get_value_after_label(line, 'federal withholding', ytd=True)
        
        if ('state tax' in line_lower or 'ga withholding' in line_lower or 'withholding' in line_lower or 'ga' in line_lower) and 'federal' not in line_lower and 'taxable' not in line_lower:
            if 'federal' not in line_lower:
                tax_val = get_value_after_label(line, 'withholding') or get_value_after_label(line, 'state')
                if tax_val:
                    result['state_tax'] = tax_val
                    result['state_name'] = 'GA'
                    result['state_tax_ytd'] = get_value_after_label(line, 'state', ytd=True) or get_value_after_label(line, 'withholding', ytd=True)
        
        if 'oasdi' in line_lower and 'taxable' not in line_lower and 'social security' not in line_lower:
            result['oasdi'] = get_value_after_label(line, 'oasdi')
            result['oasdi_ytd'] = get_value_after_label(line, 'oasdi', ytd=True)
        
        if re.search(r'medicare', line_lower) and 'taxable' not in line_lower:
            result['medicare'] = get_value_after_label(line, 'medicare')
            result['medicare_ytd'] = get_value_after_label(line, 'medicare', ytd=True)
        
        if '401k' in line_lower and 'employer' in line_lower and 'match' in line_lower:
            result['employer_match'] = get_value_after_label(line, 'match')
            result['employer_match_ytd'] = get_value_after_label(line, 'match', ytd=True)
        
        if 'hsa' in line_lower and 'employee' in line_lower:
            result['hsa'] = get_value_after_label(line, 'hsa')
            result['hsa_ytd'] = get_value_after_label(line, 'hsa', ytd=True)
        
        if 'hsa' in line_lower and 'employer' in line_lower:
            result['employer_hsa'] = get_value_after_label(line, 'hsa')
            result['employer_hsa_ytd'] = get_value_after_label(line, 'hsa', ytd=True)
        
        if 'loan repayment' in line_lower or '401k loan' in line_lower:
            result['loan_repayment'] = get_value_after_label(line, 'loan')
            result['loan_repayment_ytd'] = get_value_after_label(line, 'loan', ytd=True)
        
        if 'dependent life' in line_lower:
            result['dependent_life'] = get_value_after_label(line, 'dependent life')
            result['dependent_life_ytd'] = get_value_after_label(line, 'dependent life', ytd=True)
        
        if 'stock purchase' in line_lower or 'employee stock' in line_lower:
            result['stock_purchase'] = get_value_after_label(line, 'stock')
            result['stock_purchase_ytd'] = get_value_after_label(line, 'stock', ytd=True)
        
        if 'spousal life' in line_lower:
            result['spousal_life'] = get_value_after_label(line, 'spousal life')
            result['spousal_life_ytd'] = get_value_after_label(line, 'spousal life', ytd=True)
        
        if 'biometric credit' in line_lower and 'spousal' not in line_lower:
            result['biometric_credit'] = get_value_after_label(line, 'biometric')
            result['biometric_credit_ytd'] = get_value_after_label(line, 'biometric', ytd=True)
        
        if 'spousal biometric credit' in line_lower:
            result['spousal_biometric'] = get_value_after_label(line, 'spousal biometric')
            result['spousal_biometric_ytd'] = get_value_after_label(line, 'spousal biometric', ytd=True)
        
        if 'group term life' in line_lower:
            result['group_term_life'] = get_value_after_label(line, 'group term')
            result['group_term_life_ytd'] = get_value_after_label(line, 'group term', ytd=True)
        
        if 'floating holiday' in line_lower:
            result['floating_holiday'] = get_value_after_label(line, 'floating holiday')
            result['floating_holiday_ytd'] = get_value_after_label(line, 'floating holiday', ytd=True)
        
        if re.search(r'\bholiday\b', line_lower) and 'holiday pay' not in line_lower and 'floating' not in line_lower:
            result['holiday_pay'] = get_value_after_label(line, 'holiday')
            result['holiday_pay_ytd'] = get_value_after_label(line, 'holiday', ytd=True)
        
        if re.search(r'^vacation\s', line_lower):
            result['vacation_pay'] = get_value_after_label(line, 'vacation')
            result['vacation_pay_ytd'] = get_value_after_label(line, 'vacation', ytd=True)
        
        if re.search(r'^salary\s', line_lower):
            result['salary'] = get_value_after_label(line, 'salary')
            result['salary_ytd'] = get_value_after_label(line, 'salary', ytd=True)
        
        if 'pnc' in line_lower:
            result['bank_name'] = 'PNC Bank'
            deposit_val = get_value_after_label(line, 'pnc')
            if deposit_val:
                result['deposit_amount'] = deposit_val
        
        if 'first tech' in line_lower or 'firsttech' in line_lower:
            result['bank2_name'] = 'First Tech Federal Credit Union'
            deposit_val = get_value_after_label(line, 'first') or get_value_after_label(line, 'tech')
            if deposit_val:
                result['deposit2_amount'] = deposit_val
        
        if re.search(r'\*{6}(\d{4})', line):
            acc_match = re.search(r'\*{6}(\d{4})', line)
            if acc_match:
                if result.get('bank_name') == 'PNC Bank' and 'account_number' not in result:
                    result['account_number'] = '****' + acc_match.group(1)
                elif result.get('bank2_name') and 'account2_number' not in result:
                    result['account2_number'] = '****' + acc_match.group(1)
    
    if 'account_number' not in result:
        for line in lines:
            if re.search(r'\*+\d{4}', line):
                acc_match = re.search(r'\*+(\d{4})', line)
                if acc_match and 'account_number' not in result:
                    result['account_number'] = '****' + acc_match.group(1)
    
    return result

def extract_text_from_pdf(pdf_file):
    import io
    
    raw_text = ""
    pdf_bytes = None
    
    if hasattr(pdf_file, 'read'):
        pdf_file.seek(0)
        pdf_bytes = pdf_file.read()
        pdf_file.seek(0)
    
    if not pdf_bytes:
        print("No PDF bytes received")
        return ""
    
    print(f"PDF file size: {len(pdf_bytes)} bytes")
    print(f"PDF header: {pdf_bytes[:20]}")
    
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            print(f"pdfplumber: {len(pdf.pages)} pages")
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    raw_text += text + "\n"
        if raw_text.strip():
            print(f"pdfplumber extracted {len(raw_text)} chars")
            return raw_text
    except Exception as e:
        print(f"pdfplumber error: {e}")
    
    try:
        import fitz
        doc = fitz.open(io.BytesIO(pdf_bytes))
        print(f"PyMuPDF: {len(doc)} pages")
        for page in doc:
            text = page.get_text()
            if text:
                raw_text += text + "\n"
        doc.close()
        if raw_text.strip():
            print(f"PyMuPDF extracted {len(raw_text)} chars")
            return raw_text
    except Exception as e:
        print(f"PyMuPDF error: {e}")
    
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(io.BytesIO(pdf_bytes))
        print(f"PyPDF2: {len(reader.pages)} pages")
        for page in reader.pages:
            text = page.extract_text()
            if text:
                raw_text += text + "\n"
        if raw_text.strip():
            print(f"PyPDF2 extracted {len(raw_text)} chars")
            return raw_text
    except Exception as e:
        print(f"PyPDF2 error: {e}")
    
    print(f"Total extracted: {len(raw_text)} chars")
    return raw_text

@app.route('/import_paycheck', methods=['POST'])
def import_paycheck():
    raw_text = request.form.get('raw_text', '')
    
    pdf_file = request.files.get('pdf_file')
    
    if pdf_file and pdf_file.filename:
        if pdf_file.filename.lower().endswith('.pdf'):
            extracted = extract_text_from_pdf(pdf_file)
            if extracted.strip():
                raw_text = extracted
    
    if not raw_text or raw_text.strip() == '':
        flash('No text could be extracted. Please paste your paystub text manually below.', 'warning')
        paychecks = db.get_all_paychecks()
        next_paycheck = db.get_next_paycheck_date()
        today = datetime.now().strftime('%Y-%m-%d')
        return render_template('paychecks.html', 
                            paychecks=paychecks, 
                            next_paycheck=next_paycheck, 
                            today=today,
                            parsed_data={},
                            raw_text='',
                            active_tab='import')
    
    parsed = parse_paycheck_text(raw_text)
    
    paychecks = db.get_all_paychecks()
    next_paycheck = db.get_next_paycheck_date()
    today = datetime.now().strftime('%Y-%m-%d')
    
    return render_template('paychecks.html', 
                          paychecks=paychecks, 
                          next_paycheck=next_paycheck, 
                          today=today,
                          parsed_data=parsed,
                          raw_text=raw_text,
                          active_tab='import')

@app.route('/save_imported_paycheck', methods=['POST'])
def save_imported_paycheck():
    import json
    
    data = request.form.to_dict()
    if 'parsed_json' in data:
        try:
            parsed = json.loads(data['parsed_json'])
            data.update(parsed)
        except:
            pass
    
    numeric_fields = ['hours_worked', 'gross_pay', 'net_pay', 'pre_tax_deductions', 'employee_taxes', 
                     'post_tax_deductions', 'salary', 'biometric_credit', 'floating_holiday', 
                     'holiday_pay', 'vacation_pay', 'group_term_life', 'spousal_biometric',
                     'oasdi', 'medicare', 'federal_tax', 'state_tax', 'social_security',
                     'retirement_401k', 'add_insurance', 'dental_plan', 'eye_plan', 
                     'health_care_fsa', 'health_insurance', 'optional_life', 'hsa',
                     'loan_repayment', 'dependent_life', 'stock_purchase', 'spousal_life',
                     'employer_match', 'employer_hsa', 'deposit_amount', 'deposit2_amount',
                     'gross_pay_ytd', 'pre_tax_deductions_ytd', 'employee_taxes_ytd',
                     'post_tax_deductions_ytd', 'net_pay_ytd', 'hours_worked_ytd',
                     'salary_ytd', 'vacation_pay_ytd', 'biometric_credit_ytd', 'spousal_biometric_ytd',
                     'group_term_life_ytd', 'floating_holiday_ytd', 'holiday_pay_ytd',
                     'retirement_401k_ytd', 'health_insurance_ytd', 'dental_plan_ytd', 'eye_plan_ytd',
                     'health_care_fsa_ytd', 'optional_life_ytd', 'add_insurance_ytd', 'hsa_ytd',
                     'federal_tax_ytd', 'state_tax_ytd', 'oasdi_ytd', 'medicare_ytd',
                     'loan_repayment_ytd', 'dependent_life_ytd', 'stock_purchase_ytd', 'spousal_life_ytd',
                     'employer_match_ytd', 'employer_hsa_ytd']
    
    kwargs = {}
    for field in numeric_fields:
        val = str(data.get(field, '0')).replace(',', '')
        kwargs[field] = float(val) if val else 0
    
    kwargs['pay_date'] = data.get('check_date', '')
    kwargs['pay_period_begin'] = data.get('pay_period_begin', '')
    kwargs['pay_period_end'] = data.get('pay_period_end', '')
    kwargs['check_date'] = data.get('check_date', '')
    kwargs['check_number'] = data.get('check_number', '')
    kwargs['employee_name'] = data.get('employee_name', '')
    kwargs['employee_id'] = data.get('employee_id', '')
    kwargs['company'] = data.get('company', '')
    kwargs['state_name'] = data.get('state_name', '')
    kwargs['federal_filing_status'] = data.get('federal_filing_status', '')
    kwargs['state_filing_status'] = data.get('state_filing_status', '')
    kwargs['bank_name'] = data.get('bank_name', '')
    kwargs['account_number'] = data.get('account_number', '')
    kwargs['bank2_name'] = data.get('bank2_name', '')
    kwargs['account2_number'] = data.get('account2_number', '')
    kwargs['notes'] = f"Imported from paystub text"
    
    db.add_paycheck(**kwargs)
    flash('Paycheck imported and saved successfully!', 'success')
    return redirect(url_for('paychecks'))

@app.route('/bills')
def bills():
    filter_type = request.args.get('filter', 'all')
    all_bills = db.get_bills_with_payees()
    

    
    if filter_type == 'unpaid':
        bills = [b for b in all_bills if not b['is_paid']]
    elif filter_type == 'paid':
        bills = [b for b in all_bills if b['is_paid']]
    elif filter_type == 'overdue':
        today = datetime.now().strftime('%Y-%m-%d')
        bills = [b for b in all_bills if not b['is_paid'] and b['due_date'] < today]
    else:
        bills = all_bills
    
    # Calculate pay periods from last paycheck
    paychecks = db.get_all_paychecks()
    last_pay_date = None
    if paychecks:
        for pc in paychecks:
            if pc.get('check_date'):
                d = datetime.strptime(pc['check_date'], '%Y-%m-%d')
                if last_pay_date is None or d > last_pay_date:
                    last_pay_date = d
    
    # Generate pay periods (biweekly) from last paycheck, going back and forward
    pay_periods = []
    if last_pay_date:
        # A pay period starts the day after the previous paycheck and ends on the next paycheck
        # Go back to cover old bills
        period_start = last_pay_date + timedelta(days=1)
        for _ in range(12):
            period_start -= timedelta(weeks=2)
        
        for _ in range(26):
            period_end = period_start + timedelta(weeks=2) - timedelta(days=1)
            pay_periods.append({
                'start': period_start.strftime('%Y-%m-%d'),
                'end': period_end.strftime('%Y-%m-%d'),
                'label': f'{period_start.strftime("%m/%d/%Y")} - {period_end.strftime("%m/%d/%Y")}'
            })
            period_start = period_end + timedelta(days=1)
    
    # Group bills by month then by pay period
    months = {}
    for bill in bills:
        try:
            due = datetime.strptime(bill['due_date'], '%Y-%m-%d')
        except (ValueError, TypeError):
            continue
        month_key = due.strftime('%Y-%m')
        month_label = due.strftime('%B %Y')
        
        if month_key not in months:
            months[month_key] = {
                'label': month_label,
                'month_key': month_key,
                'periods': {},
                'total': 0
            }
        
        # Find pay period
        period_found = False
        for pp in pay_periods:
            if pp['start'] <= bill['due_date'] <= pp['end']:
                p_label = pp['label']
                if p_label not in months[month_key]['periods']:
                    months[month_key]['periods'][p_label] = {
                        'label': p_label,
                        'start': pp['start'],
                        'end': pp['end'],
                        'bills': [],
                        'total': 0
                    }
                months[month_key]['periods'][p_label]['bills'].append(bill)
                months[month_key]['periods'][p_label]['total'] += bill.get('amount', 0)
                months[month_key]['total'] += bill.get('amount', 0)
                period_found = True
                break
        
        if not period_found:
            p_label = 'Other'
            if p_label not in months[month_key]['periods']:
                months[month_key]['periods'][p_label] = {
                    'label': p_label,
                    'bills': [],
                    'total': 0
                }
            months[month_key]['periods'][p_label]['bills'].append(bill)
            months[month_key]['periods'][p_label]['total'] += bill.get('amount', 0)
            months[month_key]['total'] += bill.get('amount', 0)
    
    # Sort months chronologically
    sorted_months = sorted(months.values(), key=lambda m: m['month_key'])
    
    # Sort periods within each month by start date
    for month in sorted_months:
        periods_list = sorted(month['periods'].values(), key=lambda p: p.get('start', ''))
        month['periods_list'] = periods_list
        month['period_count'] = len(periods_list)
    
    payees = db.get_all_payees()
    current_month = datetime.now().strftime('%Y-%m')
    return render_template('bills.html', bills=bills, grouped_months=sorted_months, payees=payees, filter_type=filter_type, pay_periods=pay_periods, current_month=current_month)

@app.route('/add_bill')
def add_bill_form():
    payees = db.get_all_payees()
    return render_template('add_bill.html', payees=payees)

@app.route('/add_bill', methods=['POST'])
def add_bill():
    payee_id = request.form.get('payee_id') or None
    amount = request.form.get('amount', '0')
    due_date = request.form.get('due_date', '')
    
    if not due_date:
        flash('Due date is required!', 'danger')
        return redirect(url_for('add_bill_form'))
    
    amount = float(amount) if amount else 0.0
    
    payee_name = request.form.get('payee_name', '').strip() or None
    if not payee_id and payee_name:
        existing = db.get_payee_by_name(payee_name)
        if existing:
            payee_id = existing['id']
        else:
            payee_id = db.add_payee(payee_name, None, None, None)
    
    if not payee_name and payee_id:
        payee = db.get_payee(int(payee_id))
        if payee:
            payee_name = payee['name']
    
    category_id = request.form.get('category_id') or None
    if not category_id and payee_id:
        payee = db.get_payee(int(payee_id))
        if payee and payee.get('default_category_id'):
            category_id = payee['default_category_id']
    
    db.add_bill(
        payee_id,
        amount,
        due_date,
        1 if request.form.get('is_recurring') else 0,
        request.form.get('recurrence_type'),
        request.form.get('notes', ''),
        category_id,
        request.form.get('account') or None,
        payee_name=payee_name
    )
    flash('Bill added successfully!', 'success')
    return redirect(url_for('bills'))

@app.route('/update_bill/<int:id>', methods=['POST'])
def update_bill(id):
    payee_id = request.form['payee_id'] or None
    payee_name = None
    if payee_id:
        payee = db.get_payee(int(payee_id))
        if payee:
            payee_name = payee['name']
    db.update_bill(
        id,
        payee_id,
        float(request.form['amount']) if request.form.get('amount') else 0.0,
        request.form['due_date'],
        1 if request.form.get('is_recurring') else 0,
        request.form.get('recurrence_type'),
        request.form.get('notes', ''),
        request.form.get('category_id') or None,
        request.form.get('account') or None,
        payee_name=payee_name
    )
    flash('Bill updated successfully!', 'success')
    return redirect(url_for('bills'))

@app.route('/update_bill_ajax/<int:id>', methods=['POST'])
@app.route('/update_bill_ajax/<int:id>', methods=['POST'])
def update_bill_ajax(id):
    try:
        data = request.form
        if 'amount' in data:
            amount_str = data['amount'].strip()
            amount = float(amount_str) if amount_str else 0.0
            db.update_bill_field(id, 'amount', amount)
        if 'due_date' in data:
            db.update_bill_field(id, 'due_date', data['due_date'])
        if 'account' in data:
            db.update_bill_field(id, 'account', data['account'])
        if 'payee_name' in data:
            conn = db.get_db()
            cursor = conn.cursor()
            cursor.execute('SELECT payee_id FROM bills WHERE id = ?', (id,))
            bill = cursor.fetchone()
            payee_name = data['payee_name'].strip()
            if bill:
                if bill['payee_id']:
                    cursor.execute('UPDATE payees SET name = ? WHERE id = ?', (payee_name, bill['payee_id']))
                    cursor.execute('UPDATE bills SET payee_name = ? WHERE payee_id = ?', (payee_name, bill['payee_id']))
                else:
                    cursor.execute('INSERT INTO payees (name) VALUES (?)', (payee_name,))
                    new_id = cursor.lastrowid
                    cursor.execute('UPDATE bills SET payee_id = ?, payee_name = ? WHERE id = ?', (new_id, payee_name, id))
                conn.commit()
            conn.close()
        if 'is_paid' in data:
            is_paid = int(data['is_paid'])
            if is_paid:
                db.mark_bill_paid(id)
            else:
                db.mark_bill_unpaid(id)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/assign_bill_category/<int:id>', methods=['POST'])
def assign_bill_category(id):
    category_id = request.form.get('category_id') or None
    db.update_bill_field(id, 'category_id', category_id)
    return jsonify({'success': True})

@app.route('/mark_bill_paid/<int:id>', methods=['GET', 'POST'])
def mark_bill_paid(id):
    conn = db.get_db()
    cursor = conn.cursor()
    
    cursor.execute('SELECT is_recurring, recurrence_type, due_date FROM bills WHERE id = ?', (id,))
    bill = cursor.fetchone()
    
    if not bill:
        flash('Bill not found', 'danger')
        return redirect(url_for('bills'))
    
    is_recurring, recurrence_type, old_due_date = bill[0], bill[1], bill[2]
    
    if is_recurring and old_due_date:
        try:
            old_due = datetime.strptime(old_due_date, '%Y-%m-%d')
            today = datetime.now()
            
            new_due = old_due
            advanced = False
            while True:
                if recurrence_type == 'weekly':
                    new_due += timedelta(weeks=1)
                elif recurrence_type == 'biweekly':
                    new_due += timedelta(weeks=2)
                elif recurrence_type == 'monthly':
                    new_due = add_months(new_due, 1)
                elif recurrence_type == 'quarterly':
                    new_due = add_months(new_due, 3)
                elif recurrence_type == 'semi-monthly':
                    if new_due.day < 15:
                        new_due = new_due.replace(day=15)
                    else:
                        new_due = add_months(new_due, 1)
                        new_due = new_due.replace(day=1)
                elif recurrence_type == 'yearly':
                    new_due = add_months(new_due, 12)
                else:
                    break
                advanced = True
                if new_due > today:
                    break
            
            if advanced:
                cursor.execute('UPDATE bills SET is_paid=0, paid_date=NULL, due_date=? WHERE id=?',
                            (new_due.strftime('%Y-%m-%d'), id))
                flash(f'Bill marked as paid. Next due: {new_due.strftime("%m/%d/%Y")}', 'success')
            else:
                cursor.execute('UPDATE bills SET is_paid=1, paid_date=? WHERE id=?',
                            (datetime.now().strftime('%Y-%m-%d'), id))
                flash('Bill marked as paid!', 'success')
        except Exception as e:
            flash(f'Error updating recurring bill: {e}', 'danger')
            cursor.execute('UPDATE bills SET is_paid=1, paid_date=? WHERE id=?',
                        (datetime.now().strftime('%Y-%m-%d'), id))
    else:
        cursor.execute('UPDATE bills SET is_paid=1, paid_date=? WHERE id=?',
                    (datetime.now().strftime('%Y-%m-%d'), id))
        flash('Bill marked as paid!', 'success')
    
    conn.commit()
    
    # Store bill info for interactive budget sync (use original due_date for recurring)
    sync_date = old_due_date if (is_recurring and old_due_date) else None
    pending = session.get('_pending_budget_bills', [])
    entry = {'id': id}
    if sync_date:
        entry['due_date'] = sync_date
    if not any(e.get('id') == id for e in pending):
        pending.append(entry)
    session['_pending_budget_bills'] = pending
    
    conn.close()
    
    return redirect(url_for('bills'))

@app.route('/mark_bill_unpaid/<int:id>')
def mark_bill_unpaid(id):
    db.mark_bill_unpaid(id)
    flash('Bill marked as unpaid.', 'info')
    return redirect(url_for('bills'))

@app.route('/delete_bill/<int:id>')
def delete_bill(id):
    db.delete_bill(id)
    flash('Bill deleted.', 'info')
    next_url = request.args.get('next', url_for('bills'))
    return redirect(next_url)

@app.route('/payees')
def payees():
    payees = db.get_all_payees()
    bills = db.get_bills_with_payees()
    bills_by_payee = {}
    for bill in bills:
        pid = bill['payee_id']
        if pid not in bills_by_payee:
            bills_by_payee[pid] = []
        bills_by_payee[pid].append(bill)
    for payee in payees:
        payee['bills'] = bills_by_payee.get(payee['id'], [])
    today = datetime.now().strftime('%Y-%m-%d')
    hierarchical, budget_categories = get_hierarchical_and_flat()
    return render_template('payees.html', payees=payees, today=today, budget_categories=budget_categories)

@app.route('/export_payees_excel')
def export_payees_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    payees = db.get_all_payees()
    bills = db.get_bills_with_payees()
    bills_by_payee = {}
    for bill in bills:
        pid = bill['payee_id']
        if pid not in bills_by_payee:
            bills_by_payee[pid] = []
        bills_by_payee[pid].append(bill)
    for p in payees:
        p['bills'] = bills_by_payee.get(p['id'], [])

    wb = openpyxl.Workbook()

    # --- Sheet 1: Payees ---
    ws1 = wb.active
    ws1.title = 'Payees'
    hdr_font = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    headers1 = ['Payee Name', 'Category', 'Account #', 'Website', 'Notes', 'Total Bills', 'Total Amount', 'Next Due Date']
    for col, h in enumerate(headers1, 1):
        c = ws1.cell(row=1, column=col, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center')
        c.border = thin_border

    for row_idx, p in enumerate(payees, 2):
        total_bills = len(p['bills'])
        total_amount = sum(b['amount'] for b in p['bills']) if p['bills'] else 0
        unpaid = [b for b in p['bills'] if not b['is_paid']]
        next_due = min(b['due_date'] for b in unpaid) if unpaid else ''
        vals = [p['name'], p.get('category', '') or '', p.get('account_number', '') or '',
                p.get('website', '') or '', p.get('notes', '') or '', total_bills,
                round(total_amount, 2), next_due]
        for col, v in enumerate(vals, 1):
            c = ws1.cell(row=row_idx, column=col, value=v)
            c.border = thin_border
            if col == 1:
                c.font = Font(bold=True)

    ws1.column_dimensions['A'].width = 24
    ws1.column_dimensions['B'].width = 16
    ws1.column_dimensions['C'].width = 16
    ws1.column_dimensions['D'].width = 28
    ws1.column_dimensions['E'].width = 30
    ws1.column_dimensions['F'].width = 12
    ws1.column_dimensions['G'].width = 14
    ws1.column_dimensions['H'].width = 16

    # --- Sheet 2: Bill Details ---
    ws2 = wb.create_sheet('Bill Details')
    headers2 = ['Payee', 'Due Date', 'Amount', 'Status', 'Paid Date', 'Recurring', 'Account', 'Notes']
    for col, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center')
        c.border = thin_border

    row_idx = 2
    for bill in bills:
        status = 'Paid' if bill['is_paid'] else 'Unpaid'
        vals = [bill.get('payee_name', '') or '', bill['due_date'], bill['amount'],
                status, bill.get('paid_date', '') or '',
                'Yes' if bill['is_recurring'] else 'No',
                bill.get('account', '') or '', bill.get('notes', '') or '']
        for col, v in enumerate(vals, 1):
            c = ws2.cell(row=row_idx, column=col, value=v)
            c.border = thin_border
        row_idx += 1

    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 14
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 10
    ws2.column_dimensions['E'].width = 14
    ws2.column_dimensions['F'].width = 10
    ws2.column_dimensions['G'].width = 18
    ws2.column_dimensions['H'].width = 30

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return Response(out.getvalue(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': 'attachment; filename=payees_export.xlsx'})

@app.route('/categories')
def categories():
    payees = db.get_all_payees()
    payee_categories = list(set(p['category'] for p in payees if p.get('category')))
    budget_categories = db.get_budget_categories()
    all_categories_flat = db.get_all_budget_categories_flat()
    # Include all budget category names in the payee categories list
    budget_names = [c['name'] for c in all_categories_flat]
    combined = sorted(set(payee_categories + budget_names))
    return render_template('categories.html', payee_categories=combined, budget_categories=budget_categories, all_categories_flat=all_categories_flat)

@app.route('/add_payee_category', methods=['POST'])
def add_payee_category():
    category = request.form.get('category', '').strip()
    if category:
        db.add_payee_category_name(category)
    flash('Category added successfully!', 'success')
    return redirect(url_for('categories'))

@app.route('/delete_payee_category', methods=['POST'])
def delete_payee_category():
    category = request.form.get('category', '').strip()
    if category:
        db.delete_payee_category_by_name(category)
    flash('Category deleted.', 'info')
    return redirect(url_for('categories'))

@app.route('/rename_payee_category', methods=['POST'])
def rename_payee_category():
    old_name = request.form.get('old_name', '').strip()
    new_name = request.form.get('new_name', '').strip()
    if old_name and new_name and old_name != new_name:
        db.rename_payee_category(old_name, new_name)
        return jsonify({'success': True})
    return jsonify({'success': False, 'error': 'Invalid names'}), 400

@app.route('/update_budget_category_name/<int:id>', methods=['POST'])
def update_budget_category_name(id):
    name = request.form.get('name', '').strip()
    if not name:
        return jsonify({'success': False, 'error': 'Name is required'}), 400
    category = db.get_budget_category(id)
    if not category:
        return jsonify({'success': False, 'error': 'Category not found'}), 404
    db.update_budget_category(id, name, category['monthly_limit'], category['color'],
                              category.get('due_date', ''), category.get('notes', ''),
                              category.get('actual_spent'), category.get('parent_id'))
    return jsonify({'success': True})

@app.route('/add_payee')
def add_payee_form():
    return render_template('add_payee.html')

@app.route('/add_payee', methods=['POST'])
def add_payee():
    payee_id = db.add_payee(
        request.form['name'],
        request.form.get('category', ''),
        request.form.get('account_number', ''),
        request.form.get('notes', ''),
        request.form.get('website', '')
    )
    flash('Payee added successfully!', 'success')
    return redirect(url_for('payees'))

@app.route('/add_payee_ajax', methods=['POST'])
def add_payee_ajax():
    try:
        name = request.form.get('name', '').strip()
        if not name:
            return jsonify({'success': False, 'error': 'Name is required'}), 400
        
        payee_id = db.add_payee(
            name,
            request.form.get('category', ''),
            request.form.get('account_number', ''),
            request.form.get('notes', ''),
            request.form.get('website', '')
        )
        return jsonify({'success': True, 'payee_id': payee_id})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/match_bills_to_payees', methods=['POST'])
def match_bills_to_payees():
    try:
        conn = db.get_db()
        cursor = conn.cursor()
        cursor.execute('''UPDATE bills SET payee_id = (
            SELECT payees.id FROM payees WHERE payees.name = bills.payee_name
        ), payee_name = NULL WHERE payee_id IS NULL AND payee_name IS NOT NULL
        AND EXISTS (SELECT 1 FROM payees WHERE payees.name = bills.payee_name)''')
        matched = cursor.rowcount
        cursor.execute('UPDATE bills SET payee_id = NULL WHERE payee_id = 0')
        conn.commit()
        conn.close()
        flash(f'Matched {matched} bill(s) to existing payees!', 'success')
    except Exception as e:
        flash(f'Error matching bills: {str(e)}', 'danger')
    return redirect(url_for('payees'))

@app.route('/update_payee/<int:id>', methods=['POST'])
def update_payee(id):
    default_cat = request.form.get('default_category_id')
    default_cat = int(default_cat) if default_cat else None
    db.update_payee(
        id,
        request.form['name'],
        request.form.get('category', ''),
        request.form.get('account_number', ''),
        request.form.get('notes', ''),
        request.form.get('website', ''),
        default_category_id=default_cat
    )
    flash('Payee updated successfully!', 'success')
    return redirect(url_for('payees'))

@app.route('/delete_payee/<int:id>')
def delete_payee(id):
    db.delete_payee(id)
    flash('Payee deleted.', 'info')
    return redirect(url_for('payees'))

@app.route('/bank_accounts')
def bank_accounts():
    accounts = db.get_all_bank_accounts()
    return render_template('bank_accounts.html', accounts=accounts)

@app.route('/transactions')
def transactions():
    account_id = request.args.get('account_id', type=int)
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    search = request.args.get('search', '')
    
    txns = db.get_transactions(account_id, start_date, end_date)
    
    if search:
        txns = [t for t in txns if search.lower() in t['description'].lower()]
    
    accounts = db.get_all_bank_accounts()
    selected_account_name = None
    if account_id:
        selected_account = db.get_bank_account(account_id)
        if selected_account:
            selected_account_name = selected_account['name']
    else:
        selected_account_name = "All Accounts"
    
    return render_template('transactions.html', 
                     transactions=txns, 
                     accounts=accounts, 
                     selected_account=account_id, 
                     selected_account_name=selected_account_name)

@app.route('/import_transactions', methods=['POST'])
def import_transactions():
    account_id = request.form.get('account_id', type=int)
    clear_existing = request.form.get('clear_existing') == 'on'
    skip_duplicates = request.form.get('skip_duplicates') == 'on'
    
    if clear_existing:
        db.clear_transactions(account_id)
    
    if 'csv_file' not in request.files:
        flash('No file uploaded', 'danger')
        return redirect(url_for('transactions'))
    
    file = request.files['csv_file']
    if file.filename == '':
        flash('No file selected', 'danger')
        return redirect(url_for('transactions'))
    
    try:
        content = file.read().decode('utf-8')
        reader = csv.DictReader(io.StringIO(content))
        
        transactions_list = []
        duplicates = []
        for row in reader:
            date = row.get('Date', row.get('Posting Date', row.get('Transaction Date', ''))).strip()
            description = row.get('Description', row.get('Transaction Description', row.get('Memo', ''))).strip()
            amount_str = row.get('Amount', row.get('Transaction Amount', '0')).strip().replace('$', '').replace(',', '').replace(' ', '')
            balance_str = row.get('Balance', '0').strip().replace('$', '').replace(',', '').replace(' ', '')
            
            try:
                amount = float(amount_str)
            except:
                amount = 0
            
            try:
                balance = float(balance_str)
            except:
                balance = 0
            
            if date and amount != 0:
                transactions_list.append({
                    'date': date,
                    'description': description,
                    'amount': amount,
                    'balance': balance
                })
        
        if transactions_list:
            existing = db.get_transactions(account_id)
            existing_keys = set((tx['date'], tx['amount']) for tx in existing)
            
            new_transactions = []
            duplicate_count = 0
            for tx in transactions_list:
                key = (tx['date'], tx['amount'])
                if key in existing_keys:
                    duplicate_count += 1
                    if not skip_duplicates:
                        duplicates.append(tx)
                else:
                    new_transactions.append(tx)
            
            if duplicates and not skip_duplicates:
                return render_template('transactions.html', 
                    transactions=existing, 
                    accounts=db.get_all_bank_accounts(), 
                    selected_account=account_id,
                    duplicate_prompt=True,
                    duplicate_transactions=duplicates,
                    new_count=len(transactions_list),
                    csv_content=content)
            
            if new_transactions:
                db.add_transactions(account_id, new_transactions)
                
                # Sort by date to get the oldest (most recent) transaction for balance
                sorted_txns = sorted(new_transactions, key=lambda x: x['date'], reverse=True)
                latest_balance = sorted_txns[0]['balance'] if sorted_txns else 0
                
                account = db.get_bank_account(account_id)
                if account:
                    db.update_bank_account(account_id, account['name'], account['account_type'], 
                        account['institution'], account['account_number_last4'], latest_balance, account.get('website', ''))
                
                flash(f'Imported {len(new_transactions)} transactions', 'success')
            else:
                flash('All transactions already exist', 'info')
        else:
            flash('No transactions found in file', 'warning')
    except Exception as e:
        flash(f'Error importing: {str(e)}', 'danger')
    
    return redirect(url_for('transactions'))

@app.route('/import_transactions/confirm', methods=['POST'])
def import_transactions_confirm():
    account_id = request.form.get('account_id', type=int)
    skip_duplicates = True
    
    try:
        content = request.form.get('csv_content', '')
        if content:
            reader = csv.DictReader(io.StringIO(content))
            
            transactions_list = []
            for row in reader:
                date = row.get('Date', row.get('Posting Date', row.get('Transaction Date', ''))).strip()
                description = row.get('Description', row.get('Transaction Description', row.get('Memo', ''))).strip()
                amount_str = row.get('Amount', row.get('Transaction Amount', '0')).strip().replace('$', '').replace(',', '').replace(' ', '')
                balance_str = row.get('Balance', '0').strip().replace('$', '').replace(',', '').replace(' ', '')
                
                try:
                    amount = float(amount_str)
                except:
                    amount = 0
                
                try:
                    balance = float(balance_str)
                except:
                    balance = 0
                
                if date and amount != 0:
                    transactions_list.append({
                        'date': date,
                        'description': description,
                        'amount': amount,
                        'balance': balance
                    })
            
            existing = db.get_transactions(account_id)
            existing_keys = set((tx['date'], tx['amount']) for tx in existing)
            
            new_transactions = [tx for tx in transactions_list if (tx['date'], tx['amount']) not in existing_keys]
            
            if new_transactions:
                db.add_transactions(account_id, new_transactions)
                
                latest_balance = new_transactions[-1]['balance'] if new_transactions else 0
                if latest_balance != 0:
                    account = db.get_bank_account(account_id)
                    if account:
                        db.update_bank_account(account_id, account['name'], account['account_type'], 
                            account['institution'], account['account_number_last4'], latest_balance, account.get('website', ''))
                
                flash(f'Imported {len(new_transactions)} transactions (skipped duplicates)', 'success')
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
    
    return redirect(url_for('transactions'))
    
    file = request.files['csv_file']
    if file.filename == '':
        flash('No file selected', 'danger')
        return redirect(url_for('transactions'))
    
    try:
        content = file.read().decode('utf-8')
        reader = csv.DictReader(io.StringIO(content))
        
        transactions_list = []
        for row in reader:
            date = row.get('Date', row.get('Posting Date', row.get('Transaction Date', ''))).strip()
            description = row.get('Description', row.get('Transaction Description', row.get('Memo', ''))).strip()
            amount_str = row.get('Amount', row.get('Transaction Amount', '0')).strip().replace('$', '').replace(',', '')
            balance_str = row.get('Balance', '0').strip().replace('$', '').replace(',', '')
            
            try:
                amount = float(amount_str)
            except:
                amount = 0
            
            try:
                balance = float(balance_str)
            except:
                balance = 0
            
            if date:
                transactions_list.append({
                    'date': date,
                    'description': description,
                    'amount': amount,
                    'balance': balance
                })
        
        if transactions_list:
            db.add_transactions(account_id, transactions_list)
            flash(f'Imported {len(transactions_list)} transactions', 'success')
        else:
            flash('No transactions found in file', 'warning')
    except Exception as e:
        flash(f'Error importing: {str(e)}', 'danger')
    
    return redirect(url_for('transactions'))

@app.route('/add_bank_account')
def add_bank_account_form():
    return render_template('add_bank_account.html')

@app.route('/add_bank_account', methods=['POST'])
def add_bank_account():
    is_income = 1 if request.form.get('is_income_account') else 0
    interest_rate = 0
    try:
        interest_rate = float(request.form.get('interest_rate', 0) or 0)
    except:
        pass
    db.add_bank_account(
        request.form['name'],
        request.form['account_type'],
        request.form.get('institution', ''),
        request.form.get('account_number_last4', ''),
        float(request.form['current_balance']),
        request.form.get('website', ''),
        is_income,
        interest_rate=interest_rate
    )
    flash('Bank account added successfully!', 'success')
    return redirect(url_for('bank_accounts'))

@app.route('/update_bank_account/<int:id>', methods=['POST'])
def update_bank_account(id):
    is_income = 1 if request.form.get('is_income_account') else 0
    interest_rate = 0
    try:
        interest_rate = float(request.form.get('interest_rate', 0) or 0)
    except:
        pass
    db.update_bank_account(
        id,
        request.form['name'],
        request.form['account_type'],
        request.form.get('institution', ''),
        request.form.get('account_number_last4', ''),
        float(request.form['current_balance']),
        request.form.get('website', ''),
        is_income,
        interest_rate=interest_rate
    )
    flash('Bank account updated successfully!', 'success')
    return redirect(url_for('bank_accounts'))

@app.route('/delete_bank_account/<int:id>')
def delete_bank_account(id):
    db.delete_bank_account(id)
    flash('Bank account deleted.', 'info')
    return redirect(url_for('bank_accounts'))

@app.route('/clear_bank_account_plaid/<int:id>')
def clear_bank_account_plaid(id):
    db.clear_bank_account_plaid(id)
    flash('Bank account balance and transactions cleared. Ready for fresh Plaid sync.', 'info')
    return redirect(url_for('bank_accounts'))

@app.route('/add_transaction', methods=['POST'])
def add_transaction():
    account_id = request.form.get('account_id', type=int)
    transaction_date = request.form.get('transaction_date')
    description = request.form.get('description')
    amount = float(request.form.get('amount', 0))
    running_balance = float(request.form.get('running_balance', 0))
    
    db.add_transaction(account_id, transaction_date, description, amount, running_balance)
    
    if account_id:
        account = db.get_bank_account(account_id)
        if account:
            db.update_bank_account(account_id, account['name'], account['account_type'], 
                account['institution'], account['account_number_last4'], running_balance, account.get('website', ''))
    
    flash('Transaction added successfully!', 'success')
    return redirect(url_for('transactions', account_id=account_id))

@app.route('/update_transaction', methods=['POST'])
def update_transaction():
    id = request.form.get('id', type=int)
    account_id = request.form.get('account_id', type=int)
    transaction_date = request.form.get('transaction_date')
    description = request.form.get('description')
    amount = float(request.form.get('amount', 0))
    running_balance = float(request.form.get('running_balance', 0))
    
    db.update_transaction(id, transaction_date, description, amount, running_balance)
    
    if account_id:
        account = db.get_bank_account(account_id)
        if account:
            db.update_bank_account(account_id, account['name'], account['account_type'], 
                account['institution'], account['account_number_last4'], running_balance, account.get('website', ''))
    
    flash('Transaction updated successfully!', 'success')
    return redirect(url_for('transactions', account_id=account_id))

@app.route('/delete_transaction/<int:id>')
def delete_transaction(id):
    tx = db.get_transaction(id)
    if not tx:
        flash('Transaction not found.', 'danger')
        return redirect(url_for('transactions'))

    account_id = tx['account_id']

    # Delete first, then recompute balance from remaining transactions
    db.delete_transaction(id)

    new_balance = 0
    if account_id:
        txns = db.get_transactions(account_id)
        new_balance = txns[0]['balance'] if txns else 0
        account = db.get_bank_account(account_id)
        if account:
            db.update_bank_account(account_id, account['name'], account['account_type'], 
                account['institution'], account['account_number_last4'], new_balance, account.get('website', ''))

    flash('Transaction deleted.', 'info')
    return redirect(url_for('transactions', account_id=account_id))

@app.route('/clear_transactions')
def clear_transactions():
    account_id = request.args.get('account_id', type=int)
    if account_id:
        db.clear_transactions(account_id)
        flash('All transactions cleared for this account.', 'info')
    return redirect(url_for('transactions', account_id=account_id))

@app.route('/credit_cards')
def credit_cards():
    cards = db.get_all_credit_cards()
    return render_template('credit_cards.html', cards=cards)

@app.route('/add_credit_card')
def add_credit_card_form():
    return render_template('add_credit_card.html')

@app.route('/add_credit_card', methods=['POST'])
def add_credit_card():
    try:
        credit_limit = float(request.form.get('credit_limit', 0))
    except:
        credit_limit = 0
    try:
        current_balance = float(request.form.get('current_balance', 0))
    except:
        current_balance = 0
    try:
        interest_rate = float(request.form.get('interest_rate', 0))
    except:
        interest_rate = 0
    card_name = request.form.get('name', '').strip()
    due_date = request.form.get('due_date', '')
    card_id = db.add_credit_card(
        card_name,
        request.form.get('last_four', ''),
        credit_limit,
        current_balance,
        interest_rate,
        due_date,
        request.form.get('website', '')
    )
    # Create payee and bill for the card
    if card_name:
        payee = db.get_payee_by_name(card_name)
        payee_id = payee['id'] if payee else db.add_payee(card_name, None, None, None)
        if due_date:
            db.add_bill(payee_id, current_balance, due_date, 1, 'monthly',
                        f'Credit Card Payment - {card_name}', credit_card_id=card_id,
                        payee_name=card_name)
    flash('Credit card added successfully!', 'success')
    return redirect(url_for('credit_cards'))


@app.route('/update_credit_card/<int:id>', methods=['POST'])
def update_credit_card(id):
    try:
        credit_limit = float(request.form.get('credit_limit', 0))
    except:
        credit_limit = 0
    try:
        current_balance = float(request.form.get('current_balance', 0))
    except:
        current_balance = 0
    try:
        interest_rate = float(request.form.get('interest_rate', 0))
    except:
        interest_rate = 0
    due_date = request.form.get('due_date', '')
    db.update_credit_card(
        id,
        request.form.get('name', ''),
        request.form.get('last_four', ''),
        credit_limit,
        current_balance,
        interest_rate,
        due_date,
        request.form.get('website', '')
    )
    # Sync linked bill
    card_data = db.get_credit_card(id)
    if card_data:
        bills_data = db.get_bills_by_credit_card(id)
        for bill in bills_data:
            db.update_bill(bill['id'], bill['payee_id'], current_balance, due_date or bill['due_date'],
                          bill['is_recurring'], bill['recurrence_type'], bill.get('notes', ''),
                          payee_name=bill.get('payee_name'))
    flash('Credit card updated successfully!', 'success')
    return redirect(url_for('credit_cards'))


@app.route('/delete_credit_card/<int:id>')
def delete_credit_card(id):
    # Also delete linked bills
    for bill in db.get_bills_by_credit_card(id):
        db.delete_bill(bill['id'])
    db.delete_credit_card(id)
    flash('Credit card deleted.', 'info')
    return redirect(url_for('credit_cards'))

@app.route('/clear_credit_card_plaid/<int:id>')
def clear_credit_card_plaid(id):
    db.clear_credit_card_plaid(id)
    flash('Credit card balance cleared. Ready for fresh Plaid sync.', 'info')
    return redirect(url_for('credit_cards'))

@app.route('/budget')
def budget():
    hierarchical, categories = get_hierarchical_and_flat()
    all_bills = db.get_bills_with_payees()
    
    bills_by_category = {}
    total_actual = 0
    for cat in categories:
        cat_id = cat['id']
        cat_bills = [b for b in all_bills if b.get('category_id') == cat_id]
        bills_by_category[cat_id] = cat_bills
        manual_spent = cat.get('actual_spent') or 0
        bills_paid = sum(b['amount'] for b in cat_bills if b.get('is_paid'))
        total_actual += manual_spent + bills_paid
    
    uncategorized_bills = [b for b in all_bills if not b.get('category_id')]
    
    total_budget = sum(cat['monthly_limit'] for cat in categories)

    parent_totals = {}
    for parent in hierarchical:
        budget = parent['monthly_limit']
        manual = parent.get('actual_spent') or 0
        spent = manual + sum(b['amount'] for b in bills_by_category.get(parent['id'], []) if b.get('is_paid'))
        for child in parent.get('children', []):
            budget += child['monthly_limit']
            manual = child.get('actual_spent') or 0
            spent += manual + sum(b['amount'] for b in bills_by_category.get(child['id'], []) if b.get('is_paid'))
        parent_totals[parent['id']] = {'budget': budget, 'spent': spent}

    # Biweekly / Monthly summary at the top
    last_paychecks = db.get_all_paychecks()
    last_net = last_paychecks[0].get('net_pay', 0) if last_paychecks else 0

    biweekly_income = last_net
    biweekly_expenses = total_budget / 2 if total_budget else 0

    monthly_income = last_net * 26 / 12 if last_net else 0
    monthly_expenses = total_budget

    return render_template('budget.html', 
                        categories=categories,
                        hierarchical_categories=hierarchical,
                        bills_by_category=bills_by_category,
                        uncategorized_bills=uncategorized_bills,
                        all_bills=all_bills,
                        total_actual=total_actual,
                        total_budget=total_budget,
                        parent_totals=parent_totals,
                        biweekly_income=biweekly_income,
                        biweekly_expenses=biweekly_expenses,
                        biweekly_remaining=biweekly_income - biweekly_expenses,
                        monthly_income=monthly_income,
                        monthly_expenses=monthly_expenses,
                        monthly_remaining=monthly_income - monthly_expenses)

@app.route('/export_budget')
def export_budget():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    import io

    hierarchical, categories = get_hierarchical_and_flat()
    all_bills = db.get_bills_with_payees()

    bills_by_category = {}
    total_actual = 0
    for cat in categories:
        cat_id = cat['id']
        cat_bills = [b for b in all_bills if b.get('category_id') == cat_id]
        bills_by_category[cat_id] = cat_bills
        manual_spent = cat.get('actual_spent') or 0
        bills_paid = sum(b['amount'] for b in cat_bills if b.get('is_paid'))
        total_actual += manual_spent + bills_paid

    total_budget = sum(cat['monthly_limit'] for cat in categories)

    parent_totals = {}
    for parent in hierarchical:
        budget = parent['monthly_limit']
        manual = parent.get('actual_spent') or 0
        spent = manual + sum(b['amount'] for b in bills_by_category.get(parent['id'], []) if b.get('is_paid'))
        for child in parent.get('children', []):
            budget += child['monthly_limit']
            manual = child.get('actual_spent') or 0
            spent += manual + sum(b['amount'] for b in bills_by_category.get(child['id'], []) if b.get('is_paid'))
        parent_totals[parent['id']] = {'budget': budget, 'spent': spent}

    wb = Workbook()
    ws = wb.active
    ws.title = 'Budget Summary'

    primary_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    primary_font = Font(color='FFFFFF', bold=True, size=12)
    category_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
    child_font = Font(size=11)
    total_fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
    total_font = Font(color='FFFFFF', bold=True, size=11)
    money_fmt = '#,##0.00'
    thin_border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )

    headers = ['Category', 'Budget', 'Actual', 'Remaining']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = primary_fill
        cell.font = primary_font
        cell.alignment = Alignment(horizontal='center' if col > 1 else 'left')
        cell.border = thin_border

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

    row = 2
    parent_rows = []
    for parent in hierarchical:
        pt = parent_totals[parent['id']]
        parent_fill = PatternFill(start_color=parent['color'][1:], end_color=parent['color'][1:], fill_type='solid')
        parent_font = Font(color='FFFFFF', bold=True, size=11)

        ws.cell(row=row, column=1, value=parent['name']).font = parent_font
        ws.cell(row=row, column=1).fill = parent_fill
        ws.cell(row=row, column=2, value=round(pt['budget'], 2)).font = parent_font
        ws.cell(row=row, column=2).fill = parent_fill
        ws.cell(row=row, column=2).number_format = money_fmt
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=3, value=round(pt['spent'], 2)).font = parent_font
        ws.cell(row=row, column=3).fill = parent_fill
        ws.cell(row=row, column=3).number_format = money_fmt
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=4, value=f'=B{row}-C{row}').font = parent_font
        ws.cell(row=row, column=4).fill = parent_fill
        ws.cell(row=row, column=4).number_format = money_fmt
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = thin_border
        parent_rows.append(row)
        row += 1

        for child in parent.get('children', []):
            child_expenses = bills_by_category.get(child['id'], [])
            manual = child.get('actual_spent') or 0
            child_actual = manual + sum(b['amount'] for b in child_expenses if b.get('is_paid'))
            child_color = child.get('color', '#1565C0')

            ws.cell(row=row, column=1, value=f'  {child["name"]}')
            ws.cell(row=row, column=1).font = Font(size=11, color='333333')
            ws.cell(row=row, column=2, value=round(child['monthly_limit'], 2))
            ws.cell(row=row, column=2).number_format = money_fmt
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=3, value=round(child_actual, 2))
            ws.cell(row=row, column=3).number_format = money_fmt
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=4, value=f'=B{row}-C{row}')
            ws.cell(row=row, column=4).number_format = money_fmt
            ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')
            for c in range(1, 5):
                ws.cell(row=row, column=c).border = thin_border
            row += 1

    total_row = row
    ws.cell(row=total_row, column=1, value='Total').font = total_font
    ws.cell(row=total_row, column=1).fill = total_fill

    parent_cells_b = ','.join(f'B{r}' for r in parent_rows)
    parent_cells_c = ','.join(f'C{r}' for r in parent_rows)
    parent_cells_d = ','.join(f'D{r}' for r in parent_rows)

    ws.cell(row=total_row, column=2).value = f'=SUM({parent_cells_b})'
    ws.cell(row=total_row, column=2).font = total_font
    ws.cell(row=total_row, column=2).fill = total_fill
    ws.cell(row=total_row, column=2).number_format = money_fmt
    ws.cell(row=total_row, column=2).alignment = Alignment(horizontal='right')
    ws.cell(row=total_row, column=3).value = f'=SUM({parent_cells_c})'
    ws.cell(row=total_row, column=3).font = total_font
    ws.cell(row=total_row, column=3).fill = total_fill
    ws.cell(row=total_row, column=3).number_format = money_fmt
    ws.cell(row=total_row, column=3).alignment = Alignment(horizontal='right')
    ws.cell(row=total_row, column=4).value = f'=SUM({parent_cells_d})'
    ws.cell(row=total_row, column=4).font = total_font
    ws.cell(row=total_row, column=4).fill = total_fill
    ws.cell(row=total_row, column=4).number_format = money_fmt
    ws.cell(row=total_row, column=4).alignment = Alignment(horizontal='right')
    for c in range(1, 5):
        ws.cell(row=total_row, column=c).border = thin_border

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(output.getvalue(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': 'attachment; filename=budget_summary.xlsx'})


@app.route('/export_bills_text')
@login_required
def export_bills_text():
    bills = db.get_bills_with_payees()
    categories = db.get_all_budget_categories_flat()
    cat_map = {c['id']: c['name'] for c in categories}

    lines = []
    lines.append('=' * 60)
    lines.append('  BILLS & EXPENSES OUTLINE')
    lines.append('=' * 60)
    lines.append(f'  Generated: {datetime.now().strftime("%B %d, %Y %I:%M %p")}')
    lines.append(f'  Total Bills: {len(bills)}')
    unpaid = [b for b in bills if not b['is_paid']]
    lines.append(f'  Unpaid: {len(unpaid)}')
    lines.append('=' * 60)
    lines.append('')

    by_cat = {}
    for b in bills:
        cid = b.get('category_id')
        cat_name = cat_map.get(cid, 'Uncategorized')
        by_cat.setdefault(cat_name, []).append(b)

    for cat_name in sorted(by_cat.keys()):
        cat_bills = by_cat[cat_name]
        lines.append(f'  [{cat_name}]  ({len(cat_bills)} bills)')
        lines.append('-' * 60)
        for b in cat_bills:
            paid = 'PAID' if b['is_paid'] else 'DUE'
            amt = f"${b['amount']:.2f}"
            due = b.get('due_date') or 'N/A'
            payee = b.get('payee_name') or 'Unknown'
            acct = b.get('account') or ''
            notes = b.get('notes') or ''
            line = f'    {payee:35s} {amt:>8s}  Due: {due}  [{paid}]'
            if acct:
                line += f'  Acct: {acct}'
            lines.append(line)
            if notes:
                lines.append(f'      Notes: {notes}')
        lines.append('')

    lines.append('=' * 60)
    lines.append(f'  END  |  {len(bills)} total bills  |  {len(unpaid)} unpaid')
    lines.append('=' * 60)

    text = '\n'.join(lines)
    return Response(
        text,
        mimetype='text/plain',
        headers={'Content-Disposition': 'attachment; filename=bills_outline.txt'}
    )


@app.route('/add_budget_category')
def add_budget_category_form():
    parent_categories = db.get_all_budget_categories_flat()
    return render_template('add_budget_category.html', parent_categories=parent_categories)

@app.route('/add_budget_category', methods=['POST'])
def add_budget_category():
    parent_id = request.form.get('parent_id')
    db.add_budget_category(
        request.form['name'],
        float(request.form['monthly_limit']),
        request.form.get('color', '#2E7D32'),
        request.form.get('due_date', ''),
        request.form.get('notes', ''),
        float(request.form.get('actual_spent') or 0),
        int(parent_id) if parent_id else None
    )
    flash('Budget category added successfully!', 'success')
    return redirect(url_for('budget'))

@app.route('/update_budget_category/<int:id>', methods=['POST'])
def update_budget_category(id):
    actual = request.form.get('actual_spent')
    parent_id = request.form.get('parent_id')
    db.update_budget_category(
        id,
        request.form['name'],
        float(request.form['monthly_limit']),
        request.form.get('color', '#2E7D32'),
        request.form.get('due_date', ''),
        request.form.get('notes', ''),
        float(actual) if actual else None,
        int(parent_id) if parent_id else None
    )
    flash('Budget category updated successfully!', 'success')
    return redirect(url_for('budget'))

@app.route('/delete_budget_category/<int:id>')
def delete_budget_category(id):
    db.delete_budget_category(id)
    flash('Budget category deleted.', 'info')
    return redirect(url_for('budget'))

@app.route('/reports')
def reports():
    periods = db.get_pay_period_history()
    return render_template('reports.html', periods=periods)

@app.route('/modified_income', methods=['GET', 'POST'])
def modified_income():
    if request.method == 'POST':
        amount = float(request.form.get('amount', 0))
        entry_date = request.form.get('entry_date', datetime.now().strftime('%Y-%m-%d'))
        period_type = request.form.get('period_type', 'biweekly')
        notes = request.form.get('notes', '')
        if amount > 0:
            db.add_modified_income(amount, entry_date, period_type, notes)
            flash('Income added!', 'success')
        else:
            flash('Amount must be greater than 0', 'danger')
        return redirect(url_for('modified_income'))

    incomes = db.get_modified_incomes()
    total_bills_due = db.get_total_bills_due()
    paid_bills = db.get_paid_bills_history()
    breakdown = db.get_period_breakdown()
    total_income = sum(i['amount'] for i in incomes)
    total_paid = sum(b['amount'] for b in paid_bills)
    remaining = total_income - total_bills_due

    return render_template('modified_income.html',
                         incomes=incomes,
                         total_income=total_income,
                         total_bills_due=total_bills_due,
                         total_paid=total_paid,
                         paid_bills=paid_bills,
                         breakdown=breakdown,
                         remaining=remaining)

@app.route('/delete_modified_income/<int:id>')
def delete_modified_income(id):
    db.delete_modified_income(id)
    flash('Income entry deleted.', 'info')
    return redirect(url_for('modified_income'))

@app.route('/api/add_subcategory', methods=['POST'])
def api_add_subcategory():
    data = request.get_json()
    cat_id = db.add_budget_category(
        data['name'],
        float(data['monthly_limit']),
        data.get('color', '#2E7D32'),
        parent_id=int(data['parent_id'])
    )
    return jsonify({'success': True, 'id': cat_id})

@app.route('/api/update_category/<int:id>', methods=['POST'])
def api_update_category(id):
    data = request.get_json()
    cat = db.get_budget_category(id)
    if not cat:
        return jsonify({'error': 'Category not found'}), 404
    db.update_budget_category(
        id,
        data['name'],
        float(data['monthly_limit']),
        data.get('color', cat.get('color', '#2E7D32')),
        data.get('due_date', cat.get('due_date', '')),
        data.get('notes', cat.get('notes', '')),
        float(data.get('actual_spent', cat.get('actual_spent', 0))),
        cat.get('parent_id')
    )
    return jsonify({'success': True})


@app.route('/api/update_subcategory/<int:id>', methods=['POST'])
def api_update_subcategory(id):
    data = request.get_json()
    cat = db.get_budget_category(id)
    if not cat:
        return jsonify({'error': 'Category not found'}), 404
    db.update_budget_category(
        id,
        data['name'],
        float(data['monthly_limit']),
        data.get('color', cat.get('color', '#2E7D32')),
        actual_spent=float(data.get('actual_spent', cat.get('actual_spent', 0))),
        parent_id=cat.get('parent_id')
    )
    return jsonify({'success': True})

@app.route('/api/delete_subcategory/<int:id>', methods=['POST'])
def api_delete_subcategory(id):
    db.delete_budget_category(id)
    return jsonify({'success': True})

@app.route('/api/dashboard_stats')
def api_dashboard_stats():
    return jsonify(db.get_dashboard_stats())


@app.route('/api/account/<int:id>')
def api_account(id):
    account = db.get_bank_account(id)
    if account:
        return jsonify(dict(account))
    return jsonify({'error': 'Account not found'}), 404


# ─── Plaid Integration ─────────────────────────────────────────

def get_plaid_client():
    if not PLAID_CLIENT_ID or not PLAID_SECRET:
        return None
    from plaid.configuration import Configuration
    from plaid.api import plaid_api
    from plaid import ApiClient, Environment
    env_map = {'sandbox': Environment.Sandbox, 'development': Environment.Sandbox, 'production': Environment.Production}
    host = env_map.get(PLAID_ENV, Environment.Sandbox)
    conf = Configuration(
        host=host,
        api_key={'clientId': PLAID_CLIENT_ID, 'secret': PLAID_SECRET}
    )
    api_client = ApiClient(conf)
    return plaid_api.PlaidApi(api_client)

@app.route('/api/plaid/create_link_token', methods=['GET'])
def plaid_create_link_token():
    client = get_plaid_client()
    if not client:
        return jsonify({'error': 'Plaid not configured. Set PLAID_CLIENT_ID and PLAID_SECRET.'}), 400
    try:
        from plaid.model.link_token_create_request import LinkTokenCreateRequest
        from plaid.model.link_token_create_request_user import LinkTokenCreateRequestUser
        from plaid.model.country_code import CountryCode
        from plaid.model.products import Products
        request = LinkTokenCreateRequest(
            client_name='Budget Tracker',
            language='en',
            country_codes=[CountryCode('US')],
            user=LinkTokenCreateRequestUser(client_user_id='user-1'),
            products=[Products('transactions')]
        )
        response = client.link_token_create(request)
        return jsonify({'link_token': response.to_dict()['link_token']})
    except Exception as e:
        app.logger.error('Plaid link_token_create error: %s', str(e))
        return jsonify({'error': str(e)}), 500

@app.route('/api/plaid/exchange_public_token', methods=['POST'])
def plaid_exchange_public_token():
    client = get_plaid_client()
    if not client:
        return jsonify({'error': 'Plaid not configured'}), 400
    try:
        from plaid.model.item_public_token_exchange_request import ItemPublicTokenExchangeRequest
        public_token = request.json.get('public_token')
        exchange_req = ItemPublicTokenExchangeRequest(public_token=public_token)
        exchange_res = client.item_public_token_exchange(exchange_req)
        access_token = exchange_res.to_dict()['access_token']
        item_id = exchange_res.to_dict()['item_id']

        from plaid.model.accounts_get_request import AccountsGetRequest
        accounts_req = AccountsGetRequest(access_token=access_token)
        accounts_res = client.accounts_get(accounts_req)
        accounts_data = accounts_res.to_dict()['accounts']

        institution_name = request.json.get('institution', '')
        item_pk = db.add_plaid_item(access_token, item_id, institution_name)

        created = []
        for acct in accounts_data:
            acct_type = acct.get('type', '')
            acct_subtype = acct.get('subtype', '')
            acct_data = {
                'plaid_account_id': acct['account_id'],
                'name': acct['name'],
                'mask': acct.get('mask', ''),
                'balances': acct.get('balances', {}),
                'type': acct_type,
                'subtype': acct_subtype
            }
            if acct_type == 'depository':
                balance = acct_data['balances'].get('current', 0) or 0
                db.add_bank_account(
                    name=acct_data['name'],
                    account_type=acct_subtype or 'checking',
                    institution=institution_name,
                    account_number_last4=acct_data['mask'],
                    current_balance=balance,
                    website='',
                    plaid_account_id=acct_data['plaid_account_id'],
                    plaid_item_id=item_pk
                )
                created.append({'type': 'bank', 'name': acct_data['name']})
            elif acct_type == 'investment':
                balance = acct_data['balances'].get('current', 0) or 0
                db.add_bank_account(
                    name=acct_data['name'],
                    account_type='investment',
                    institution=institution_name,
                    account_number_last4=acct_data['mask'],
                    current_balance=balance,
                    website='',
                    plaid_account_id=acct_data['plaid_account_id'],
                    plaid_item_id=item_pk
                )
                created.append({'type': 'bank', 'name': acct_data['name']})
            elif acct_type == 'credit':
                balance = acct_data['balances'].get('current', 0) or 0
                limit = acct_data['balances'].get('limit', 0) or 0
                card_name = acct_data['name']
                card_id = db.add_credit_card(
                    name=card_name,
                    last_four=acct_data['mask'],
                    credit_limit=limit,
                    current_balance=balance,
                    interest_rate=0,
                    due_date='',
                    plaid_account_id=acct_data['plaid_account_id'],
                    plaid_item_id=item_pk
                )
                if card_name:
                    payee = db.get_payee_by_name(card_name)
                    payee_id = payee['id'] if payee else db.add_payee(card_name, None, None, None)
                    db.add_bill(payee_id, balance, '', 1, 'monthly',
                                f'Credit Card Payment - {card_name}', credit_card_id=card_id,
                                payee_name=card_name)
                created.append({'type': 'credit', 'name': card_name})
            elif acct_type == 'loan':
                balance = acct_data['balances'].get('current', 0) or 0
                db.add_bank_account(
                    name=acct_data['name'],
                    account_type='loan',
                    institution=institution_name,
                    account_number_last4=acct_data['mask'],
                    current_balance=balance,
                    website='',
                    plaid_account_id=acct_data['plaid_account_id'],
                    plaid_item_id=item_pk
                )
                created.append({'type': 'bank', 'name': acct_data['name']})

        try:
            from plaid.model.transactions_sync_request import TransactionsSyncRequest
            cursor_val = ''
            has_more = True
            while has_more:
                sync_req = TransactionsSyncRequest(access_token=access_token, cursor=cursor_val)
                sync_res = client.transactions_sync(sync_req)
                sync_data = sync_res.to_dict()
                accounts_map = {a['plaid_account_id']: a for a in db.get_accounts_by_plaid_item(item_pk)}
                for tx in sync_data.get('added', []):
                    local = accounts_map.get(tx['account_id'])
                    if not local:
                        continue
                    plaid_amount = tx.get('amount', 0) or 0
                    date = tx.get('date', '')
                    desc = tx.get('merchant_name') or tx.get('name') or ''
                    db.upsert_plaid_transaction(
                        local_account_id=local['id'], plaid_tx_id=tx['transaction_id'],
                        date=date, description=desc, amount=-plaid_amount, running_balance=0
                    )
                for tx in sync_data.get('modified', []):
                    local = accounts_map.get(tx['account_id'])
                    if not local:
                        continue
                    plaid_amount = tx.get('amount', 0) or 0
                    date = tx.get('date', '')
                    desc = tx.get('merchant_name') or tx.get('name') or ''
                    db.upsert_plaid_transaction(
                        local_account_id=local['id'], plaid_tx_id=tx['transaction_id'],
                        date=date, description=desc, amount=-plaid_amount, running_balance=0
                    )
                for tx in sync_data.get('removed', []):
                    db.delete_plaid_transaction(tx['transaction_id'])
                cursor_val = sync_data.get('next_cursor', '')
                has_more = sync_data.get('has_more', False)
            if cursor_val:
                db.update_plaid_cursor(item_pk, cursor_val)
        except Exception:
            pass

        return jsonify({'success': True, 'accounts': created})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/plaid/sync_balances', methods=['POST'])
def plaid_sync_balances():
    client = get_plaid_client()
    if not client:
        return jsonify({'error': 'Plaid not configured'}), 400
    try:
        from plaid.model.accounts_balance_get_request import AccountsBalanceGetRequest
        items = db.get_plaid_items()
        results = []
        for item in items:
            try:
                req = AccountsBalanceGetRequest(access_token=item['access_token'])
                res = client.accounts_balance_get(req)
                accounts = res.to_dict()['accounts']
                for acct in accounts:
                    aid = acct['account_id']
                    balances = acct.get('balances', {}) or {}
                    balance = balances.get('current', 0) or 0
                    limit_val = balances.get('limit', 0) or 0
                    conn = db.get_db()
                    cursor = conn.cursor()
                    cursor.execute('UPDATE bank_accounts SET current_balance=? WHERE plaid_account_id=?', (balance, aid))
                    if cursor.rowcount == 0:
                        cursor.execute('UPDATE credit_cards SET current_balance=?, credit_limit=? WHERE plaid_account_id=?', (balance, limit_val, aid))
                    conn.commit()
                    conn.close()
                results.append({'item': item.get('institution_name', ''), 'status': 'ok'})
            except Exception as e:
                results.append({'item': item.get('institution_name', ''), 'status': 'error', 'error': str(e)})
        return jsonify({'success': True, 'results': results})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/plaid/sync_transactions', methods=['POST'])
def plaid_sync_transactions():
    client = get_plaid_client()
    if not client:
        return jsonify({'error': 'Plaid not configured'}), 400
    try:
        from plaid.model.transactions_sync_request import TransactionsSyncRequest
        items = db.get_plaid_items()
        totals = {'added': 0, 'modified': 0, 'removed': 0}
        for item in items:
            cursor_val = item.get('plaid_cursor') or ''
            has_more = True
            while has_more:
                req = TransactionsSyncRequest(access_token=item['access_token'], cursor=cursor_val)
                res = client.transactions_sync(req)
                data = res.to_dict()
                accounts_map = {a['plaid_account_id']: a for a in db.get_accounts_by_plaid_item(item['id'])}
                for tx in data.get('added', []):
                    local = accounts_map.get(tx['account_id'])
                    if not local:
                        continue
                    plaid_amount = tx.get('amount', 0) or 0
                    date = tx.get('date', '')
                    desc = tx.get('merchant_name') or tx.get('name') or ''
                    db.upsert_plaid_transaction(
                        local_account_id=local['id'],
                        plaid_tx_id=tx['transaction_id'],
                        date=date,
                        description=desc,
                        amount=-plaid_amount,
                        running_balance=0
                    )
                    totals['added'] += 1
                for tx in data.get('modified', []):
                    local = accounts_map.get(tx['account_id'])
                    if not local:
                        continue
                    plaid_amount = tx.get('amount', 0) or 0
                    date = tx.get('date', '')
                    desc = tx.get('merchant_name') or tx.get('name') or ''
                    db.upsert_plaid_transaction(
                        local_account_id=local['id'],
                        plaid_tx_id=tx['transaction_id'],
                        date=date,
                        description=desc,
                        amount=-plaid_amount,
                        running_balance=0
                    )
                    totals['modified'] += 1
                for tx in data.get('removed', []):
                    db.delete_plaid_transaction(tx['transaction_id'])
                    totals['removed'] += 1
                cursor_val = data.get('next_cursor', '')
                has_more = data.get('has_more', False)
            if cursor_val:
                db.update_plaid_cursor(item['id'], cursor_val)
        msg = f"Synced: {totals['added']} added, {totals['modified']} modified, {totals['removed']} removed"
        return jsonify({'success': True, 'message': msg, 'totals': totals})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/plaid_settings', methods=['GET', 'POST'])
def plaid_settings():
    if request.method == 'POST':
        cfg = {
            'client_id': request.form.get('client_id', '').strip(),
            'secret': request.form.get('secret', '').strip(),
            'environment': request.form.get('environment', 'sandbox')
        }
        with open(plaid_cfg_file, 'w') as f:
            json.dump(cfg, f)
        os.chmod(plaid_cfg_file, 0o600)
        # Reload config without restart
        global PLAID_CLIENT_ID, PLAID_SECRET, PLAID_ENV
        PLAID_CLIENT_ID = cfg['client_id']
        PLAID_SECRET = cfg['secret']
        PLAID_ENV = cfg['environment']
        flash('Plaid credentials saved!', 'success')
        return redirect(url_for('bank_accounts'))
    return render_template('plaid_settings.html',
                          client_id=PLAID_CLIENT_ID,
                          env=PLAID_ENV)


@app.route('/import_paycheck_pdf', methods=['POST'])
def import_paycheck_pdf():
    if 'pdf_file' not in request.files:
        flash('No file uploaded', 'danger')
        return redirect(url_for('import_paystub'))
    file = request.files['pdf_file']
    if file.filename == '':
        flash('No file selected', 'danger')
        return redirect(url_for('import_paystub'))
    if file and file.filename.endswith('.pdf'):
        import os
        import uuid
        import tempfile
        
        try:
            from pypdf import PdfReader
            
            file_content = file.read()
            with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
                tmp.write(file_content)
                tmp_path = tmp.name
            
            text = PdfReader(tmp_path).pages[0].extract_text()
            os.unlink(tmp_path)
        except Exception as e:
            flash(f'Error reading PDF: {str(e)}', 'danger')
            return redirect(url_for('import_paystub'))
        
        data = parse_paycheck_text(text)
        import json
        return render_template('import_paystub.html', parsed_data=data)
    
    flash('Invalid file type', 'danger')
    return redirect(url_for('import_paystub'))


@app.route('/import_paystub')
def import_paystub_form():
    return render_template('import_paystub.html')


@app.route('/view_paycheck/<int:id>')
def view_paycheck(id):
    paycheck = db.get_paycheck(id)
    if not paycheck:
        flash('Paycheck not found', 'danger')
        return redirect(url_for('paychecks'))
    return render_template('view_paycheck.html', paycheck=paycheck)



@app.route('/import_statement', methods=['GET', 'POST'])
def import_statement():
    if request.method == 'POST':
        account_id = request.form.get('account_id', type=int)
        new_name = request.form.get('new_name', '').strip()
        new_type = request.form.get('new_type', 'checking')
        new_institution = request.form.get('new_institution', '').strip()
        new_last4 = request.form.get('new_last4', '').strip()
        new_balance = request.form.get('new_balance', type=float) or 0
        skip_duplicates = request.form.get('skip_duplicates') == 'on'

        if 'csv_file' not in request.files:
            flash('No file uploaded', 'danger')
            return redirect(url_for('import_statement'))

        file = request.files['csv_file']
        if file.filename == '':
            flash('No file selected', 'danger')
            return redirect(url_for('import_statement'))

        try:
            raw = file.read()
            fname = file.filename.lower()

            if fname.endswith('.pdf'):
                import pdfplumber
                pdf_file = io.BytesIO(raw)
                text_lines = []
                with pdfplumber.open(pdf_file) as pdf:
                    for page in pdf.pages:
                        page_text = page.extract_text()
                        if page_text:
                            text_lines.extend(page_text.split('\n'))

                transactions_list = []
                date_pattern = re.compile(r'(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})')
                for line in text_lines:
                    line = line.strip()
                    if not line:
                        continue
                    dm = date_pattern.search(line)
                    if not dm:
                        continue
                    amounts = re.findall(r'[-]?\$?[\d,]+\.\d{2}', line)
                    if not amounts:
                        continue
                    amounts = [a.replace('$', '').replace(',', '') for a in amounts]
                    nums = []
                    for a in amounts:
                        try:
                            nums.append(float(a))
                        except:
                            pass
                    if not nums:
                        continue
                    # If two amounts: last is likely running balance
                    amount = nums[0]
                    balance = nums[-1] if len(nums) > 1 else 0
                    # Determine sign: if line has "CREDIT" or positive context
                    if 'credit' in line.lower() or 'deposit' in line.lower():
                        if amount < 0:
                            amount = abs(amount)
                    elif 'debit' in line.lower() or 'withdrawal' in line.lower():
                        if amount > 0:
                            amount = -amount
                    desc = line[:dm.start()].strip() + ' ' + line[dm.end():].strip()
                    for a in amounts:
                        desc = desc.replace(a, '').replace('$', '').strip()
                    desc = re.sub(r'\s+', ' ', desc).strip()
                    date_raw = dm.group(1).replace('-', '/')
                    parts = date_raw.split('/')
                    if len(parts[0]) == 4:
                        date_raw = f'{parts[1]}/{parts[2]}/{parts[0]}'
                    transactions_list.append({
                        'date': date_raw,
                        'description': desc[:100] or 'Unknown',
                        'amount': amount,
                        'balance': balance
                    })

                if not transactions_list:
                    flash('No transactions found in PDF. Try CSV format.', 'warning')
                    return redirect(url_for('import_statement'))
            else:
                content = raw.decode('utf-8')
                reader = csv.DictReader(io.StringIO(content))

                transactions_list = []
                for row in reader:
                    date = row.get('Date', row.get('Posting Date', row.get('Transaction Date', ''))).strip()
                    description = row.get('Description', row.get('Transaction Description', row.get('Memo', ''))).strip()
                    amount_str = row.get('Amount', row.get('Transaction Amount', '0')).strip().replace('$', '').replace(',', '').replace(' ', '')
                    balance_str = row.get('Balance', '0').strip().replace('$', '').replace(',', '').replace(' ', '')

                    try:
                        amount = float(amount_str)
                    except:
                        amount = 0

                    try:
                        balance = float(balance_str)
                    except:
                        balance = 0

                    if date and amount != 0:
                        transactions_list.append({
                            'date': date,
                            'description': description,
                            'amount': amount,
                            'balance': balance
                        })

                if not transactions_list:
                    flash('No transactions found in file', 'warning')
                    return redirect(url_for('import_statement'))

            # Determine target account
            target_account_id = account_id
            if not target_account_id and new_name:
                db.add_bank_account(new_name, new_type, new_institution, new_last4, new_balance)
                accounts = db.get_all_bank_accounts()
                target_account_id = accounts[-1]['id'] if accounts else None

            if not target_account_id:
                flash('Select an existing account or fill in new account details', 'danger')
                return redirect(url_for('import_statement'))

            # Import transactions
            existing = db.get_transactions(target_account_id)
            existing_keys = set((tx['date'], tx['amount']) for tx in existing)

            new_transactions = []
            dup_count = 0
            for tx in transactions_list:
                key = (tx['date'], tx['amount'])
                if key in existing_keys:
                    dup_count += 1
                    if not skip_duplicates:
                        continue
                else:
                    new_transactions.append(tx)

            if new_transactions:
                db.add_transactions(target_account_id, new_transactions)

                # Update balance from latest transaction
                sorted_txns = sorted(transactions_list, key=lambda x: x['date'], reverse=True)
                latest_balance = sorted_txns[0]['balance'] if sorted_txns else 0
                account = db.get_bank_account(target_account_id)
                if account:
                    db.update_bank_account(target_account_id, account['name'],
                        account['account_type'], account['institution'],
                        account['account_number_last4'], latest_balance, account.get('website', ''))

                msg = f'Imported {len(new_transactions)} transactions'
                if dup_count:
                    msg += f' ({dup_count} duplicates skipped)'
                flash(msg, 'success')
            else:
                flash('All transactions already exist', 'info')

        except Exception as e:
            flash(f'Error importing: {str(e)}', 'danger')

        return redirect(url_for('bank_accounts'))

    accounts = db.get_all_bank_accounts()
    return render_template('import_statement.html', accounts=accounts)


@app.route('/interactive_budget')
def interactive_budget():
    payees = db.get_all_payees()
    payee_names = sorted(set(p['name'] for p in payees if p['name'].strip()))
    hierarchical, _ = get_hierarchical_and_flat()
    bills = db.get_bills_with_payees()
    pending_sync = session.pop('_pending_budget_bills', [])
    last_paychecks = db.get_all_paychecks()
    biweekly_income = last_paychecks[0].get('net_pay', 0) if last_paychecks else 0
    return render_template('interactive_budget.html', payees=payee_names, budget_categories=hierarchical, bills=bills, pending_budget_sync=pending_sync, biweekly_income=biweekly_income)


@app.route('/goals')
def goals():
    get_category_options()
    accounts = db.get_all_bank_accounts()
    assets = [a for a in accounts if a['account_type'] in ('checking','savings','money_market','investment')]
    loans = [a for a in accounts if a['account_type'] == 'loan']
    credit_cards = db.get_all_credit_cards()
    total_assets = sum(a['current_balance'] for a in assets)
    total_loans = sum(a['current_balance'] for a in loans)
    total_credit = sum(c['current_balance'] for c in credit_cards)
    last_paychecks = db.get_all_paychecks()
    last_net = last_paychecks[0].get('net_pay', 0) if last_paychecks else 0
    monthly_income = last_net * 26 / 12 if last_net else 0
    monthly_budget = db._cached_total_budget or 0
    all_debts = []
    for c in credit_cards:
        all_debts.append({'name': c['name'], 'type': 'Credit Card', 'balance': c['current_balance'], 'rate': c.get('interest_rate', 0) or 0, 'id': 'cc_' + str(c['id'])})
    for l in loans:
        all_debts.append({'name': l['name'], 'type': 'Loan', 'balance': l['current_balance'], 'rate': l.get('interest_rate', 0) or 0, 'id': 'loan_' + str(l['id'])})

    return render_template('goals.html',
        assets=assets, loans=loans, credit_cards=credit_cards,
        total_assets=total_assets, total_loans=total_loans,
        total_credit=total_credit,
        total_debt=total_loans + total_credit,
        net_worth=total_assets - total_loans - total_credit,
        last_net=last_net, monthly_income=monthly_income,
        monthly_budget=monthly_budget,
        all_debts=all_debts)


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', '8080')), debug=False)
